import subprocess
import sys
import re
from openpyxl import load_workbook

def find_signature(target_dump, extension):
    # find signature
    db = load_workbook("./Database/Database_v2.xlsx", data_only=True)
    db_sheet = db['Sheet1']
    
    signature = []
    for i in range(2, 77):
        if str(db_sheet.cell(i, 1).value) == str(extension):
            signature.append(db_sheet.cell(i, 2).value)
    if not signature:
        print("Don't have signature in DB")
        exit(1)
    
    print(signature)
    return signature


def command_dd(img, i, cnt):
    dd = 'dd if=' + img + ' bs=4096 count=12 skip=' + block_list[i] + ' > recover' + str(cnt)
    p = subprocess.run(dd, shell=True, universal_newlines=True)


if __name__ == "__main__":

    target_dump = sys.argv[1]
    if len(sys.argv) is 2:
        print('python3 [filename] [dumpfile] [signature]')
        sys.exit(1)


    extension = sys.argv[2].upper()
    signature = find_signature(target_dump, extension)
    
    for sig in signature:
        sigfind = 'sigfind -b 4096 ' + sig + ' ' + target_dump + '; exit 0'
        block_list = subprocess.check_output(sigfind, shell=True, universal_newlines=True)
        list_output = block_list.split(' ')
        list_output = list_output[9:]
        if not list_output:
            continue
        block_list = ' '.join(list_output).split(' ')
        cnt=1
        for i in range(0, len(block_list), 2):
            command_dd(target_dump, i, cnt)
            cnt+=1
    
