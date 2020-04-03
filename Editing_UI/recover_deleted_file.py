import subprocess
import sys
import re
from openpyxl import load_workbook

class recover_file():
	header = []
	dump_path = []
	recovery_path = []
	block_list = []

	def __init__(self, sigHeader="jpeg", d_path="../04_dump.img", r_path="../"):
		self.header = sigHeader
		self.dump_path = d_path
		self.recovery_path = r_path


	def command_dd(self, img, i, cnt):
		dd = 'dd if=' + img + ' bs=4096 count=12 skip=' + self.block_list[i] + ' > ' + self.recovery_path + 'recover' + str(cnt)
		p = subprocess.run(dd, shell=True, universal_newlines=True)


	def find_signature(self, target_dump, extension):
		# find signature
		db = load_workbook("../Database/Database_v2.xlsx", data_only=True)
		db_sheet = db['Sheet1']
		
		signature = []
		for i in range(2, 77):
			if str(db_sheet.cell(i, 1).value) == str(extension):
				signature.append(db_sheet.cell(i, 2).value)
			if not signature:
				print("Don't have signature in DB")
				#exit(1)
		print(signature)
		print("found signature")
		return signature


	def start_recovery(self):
		target_dump = self.dump_path
		#if len(sys.argv) is 2:
		#	print('python3 [filename] [dumpfile] [signature]')
		#	sys.exit(1)		

		extension = self.header.upper()
		signature = self.find_signature(target_dump, extension)
		print('gonna start')
		for sig in signature:
			sigfind = 'sigfind -b 4096 ' + sig + ' ' + target_dump + '; exit 0'
			self.block_list = subprocess.check_output(sigfind, shell=True, universal_newlines=True)
			list_output = self.block_list.split(' ')
			list_output = list_output[9:]
			if not list_output:
				continue
			self.block_list = ' '.join(list_output).split(' ')
			cnt=1
			for i in range(0, len(self.block_list), 2):
				self.command_dd(target_dump, i, cnt)
				cnt+=1
			fd = open("recover_list.tmp", 'a+', encoding = 'utf-8')
			fd.write('.' + self.header + ' files -' + str(cnt))
			fd.close()
